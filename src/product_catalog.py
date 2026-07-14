#!/usr/bin/env python3
"""Durable multi-product catalog import and retrieval for Admira IA."""

from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import shutil
import unicodedata
from pathlib import Path

from codex_brand_guides import (
    MAX_BASE_PRODUCTS,
    MAX_PRODUCT_GUIDES,
    PRODUCT_DIR,
    ROOT_DIR,
    guide_library,
    normalize_product_payload,
    product_fields,
    product_slug,
    read_text,
    refresh_offer_map,
    save_product_guide,
)


PRODUCT_IMPORT_DIR = ROOT_DIR / "dashboard" / "data" / "product-imports"
CATALOG_INDEX_FILE = ROOT_DIR / "brand_guides" / "catalog-index.json"
MAX_PRODUCT_IMPORT_BYTES = 50 * 1024 * 1024
SUPPORTED_CATALOG_EXTENSIONS = {".csv", ".tsv", ".xlsx", ".xls", ".pdf", ".json"}


HEADER_ALIASES = {
    "name": ("nombre", "producto", "nombre producto", "product", "product name", "offer", "oferta", "servicio"),
    "sku": ("sku", "codigo", "código", "referencia", "product id", "id producto", "id"),
    "kind": ("tipo", "tipo producto", "product type", "tipo oferta", "clase"),
    "category": ("categoria", "categoría", "category", "coleccion", "colección"),
    "status": ("estado", "status", "activo", "active"),
    "url": ("url", "link", "enlace", "website", "landing", "pagina", "página"),
    "price": ("precio", "price", "rango precio", "precio venta"),
    "cost": ("costo", "cost", "coste", "costo unitario"),
    "margin": ("margen", "margin", "margen bruto"),
    "short_description": ("resumen", "descripcion corta", "descripción corta", "short description"),
    "description": ("descripcion", "descripción", "description", "detalle", "descripcion detallada", "descripción detallada"),
    "includes": ("incluye", "que incluye", "qué incluye", "inclusions", "contenido"),
    "features": ("caracteristicas", "características", "features", "atributos", "especificaciones"),
    "variants": ("variantes", "variants", "opciones", "tallas", "sabores"),
    "availability": ("disponibilidad", "availability", "stock", "inventario"),
    "audience": ("audiencia", "publico", "público", "cliente ideal", "target audience", "para quien"),
    "not_for": ("para quien no", "no recomendado", "not for"),
    "pain": ("dolor", "problema", "necesidad", "pain", "problem"),
    "desire": ("beneficio", "deseo", "resultado", "benefit", "value proposition", "promesa"),
    "objections": ("objeciones", "objections", "dudas frecuentes"),
    "show": ("mostrar", "debe mostrar", "must show"),
    "avoid": ("evitar", "no mostrar", "must avoid"),
    "assets": ("fotos", "imagenes", "imágenes", "assets", "media", "image urls"),
    "tags": ("etiquetas", "tags", "keywords", "palabras clave"),
    "components": ("componentes", "productos incluidos", "bundle products", "conjunto", "pack"),
    "cross_sell": ("venta cruzada", "cross sell", "cross-sell", "productos relacionados"),
    "upsell": ("upsell", "upgrade", "mejora sugerida"),
    "additional_details": ("detalles adicionales", "otros datos", "additional details", "notas"),
}


def _normalized(value):
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", " ", text.lower()).strip()
    return re.sub(r"\s+", " ", text)


HEADER_LOOKUP = {
    _normalized(alias): canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _cell_text(value):
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Sí" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _join_extra_fields(extra):
    return " | ".join(f"{key}: {value}" for key, value in extra.items() if value)


def rows_to_products(headers, rows, source=""):
    mapped_headers = []
    for index, raw_header in enumerate(headers or []):
        label = _cell_text(raw_header) or f"columna_{index + 1}"
        mapped_headers.append((label, HEADER_LOOKUP.get(_normalized(label), "")))
    products = []
    rejected = []
    for row_number, row in enumerate(rows or [], start=2):
        product = {}
        extras = {}
        for index, (label, canonical) in enumerate(mapped_headers):
            value = _cell_text(row[index] if index < len(row) else "")
            if not value:
                continue
            if canonical:
                previous = str(product.get(canonical) or "").strip()
                product[canonical] = f"{previous} | {value}" if previous and value not in previous else value
            else:
                extras[label] = value
        if extras:
            product["additional_details"] = _join_extra_fields(extras)
        if source:
            product["source"] = source
        if not str(product.get("name") or "").strip():
            if any(str(value or "").strip() for value in product.values()):
                rejected.append({"row": row_number, "reason": "missing_product_name"})
            continue
        products.append(product)
    return products, rejected


def _extract_delimited(path):
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:8192]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel_tab if path.suffix.lower() == ".tsv" else csv.excel
    rows = list(csv.reader(raw.splitlines(), dialect))
    if not rows:
        return [], []
    return rows_to_products(rows[0], rows[1:], source=path.name)


def _extract_xlsx(path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("La lectura de Excel no está instalada en esta versión.") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    products = []
    rejected = []
    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue
        sheet_products, sheet_rejected = rows_to_products(rows[0], rows[1:], source=f"{path.name} / {sheet.title}")
        products.extend(sheet_products)
        rejected.extend({**item, "sheet": sheet.title} for item in sheet_rejected)
    workbook.close()
    return products, rejected


def _extract_xls(path):
    try:
        import xlrd
    except ImportError as exc:
        raise ValueError("Para leer Excel .xls instala xlrd o guarda el archivo como .xlsx.") from exc
    workbook = xlrd.open_workbook(path)
    products = []
    rejected = []
    for sheet in workbook.sheets():
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        if not rows:
            continue
        sheet_products, sheet_rejected = rows_to_products(rows[0], rows[1:], source=f"{path.name} / {sheet.name}")
        products.extend(sheet_products)
        rejected.extend({**item, "sheet": sheet.name} for item in sheet_rejected)
    return products, rejected


def _extract_json(path):
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict):
        payload = payload.get("products") or payload.get("productos") or [payload]
    if not isinstance(payload, list):
        raise ValueError("El JSON debe contener una lista de productos.")
    products = [dict(item) for item in payload if isinstance(item, dict)]
    for item in products:
        item.setdefault("source", path.name)
    return products, []


def _pdf_key_value_products(text, source):
    products = []
    current = None
    for raw_line in str(text or "").splitlines():
        line = re.sub(r"^[\s•*\-]+", "", raw_line).strip()
        if not line:
            continue
        match = re.match(r"^(?:producto|nombre(?: del producto)?|product(?: name)?)\s*[:\-]\s*(.+)$", line, re.I)
        if match:
            if current and current.get("name"):
                products.append(current)
            current = {"name": match.group(1).strip(), "source": source}
            continue
        if current is None or ":" not in line:
            continue
        key, value = line.split(":", 1)
        canonical = HEADER_LOOKUP.get(_normalized(key))
        value = value.strip()
        if canonical and value:
            current[canonical] = value
        elif value:
            previous = str(current.get("additional_details") or "")
            current["additional_details"] = f"{previous} | {key.strip()}: {value}".strip(" |")
    if current and current.get("name"):
        products.append(current)
    return products


def _extract_pdf(path):
    try:
        from pypdf import PdfReader
    except ImportError as exc:
        raise ValueError("La lectura de PDF no está instalada en esta versión.") from exc
    reader = PdfReader(str(path))
    pages = [str(page.extract_text() or "").strip() for page in reader.pages]
    text = "\n\n".join(page for page in pages if page).strip()
    if not text:
        return [], [], "", "El PDF no contiene texto seleccionable; envía una versión con texto o una hoja Excel."
    products = _pdf_key_value_products(text, path.name)
    return products, [], text, "" if products else "El PDF fue guardado y leído, pero necesita que el agente estructure sus productos desde el texto extraído."


def _allowed_document_roots():
    candidates = [
        PRODUCT_IMPORT_DIR,
        ROOT_DIR / "dashboard" / "data" / "hermes-home" / "cache" / "documents",
        ROOT_DIR / "dashboard" / "data" / "hermes-workspace" / "current" / "uploads",
        ROOT_DIR / "output",
        Path.home() / ".hermes" / "cache" / "documents",
    ]
    configured = str(os.environ.get("HERMES_HOME") or "").strip()
    if configured:
        candidates.append(Path(configured).expanduser() / "cache" / "documents")
    return [path.expanduser().resolve(strict=False) for path in candidates]


def safe_catalog_document_path(raw_path):
    try:
        path = Path(str(raw_path or "")).expanduser().resolve(strict=True)
    except (OSError, RuntimeError, ValueError) as exc:
        raise ValueError("No encontré el archivo de productos.") from exc
    if not path.is_file() or path.suffix.lower() not in SUPPORTED_CATALOG_EXTENSIONS:
        raise ValueError("Admite catálogos PDF, Excel, CSV, TSV o JSON.")
    if path.stat().st_size > MAX_PRODUCT_IMPORT_BYTES:
        raise ValueError("El catálogo supera el límite de 50 MB.")
    for root in _allowed_document_roots():
        try:
            path.relative_to(root)
            return path
        except ValueError:
            continue
    raise ValueError("Ese archivo está fuera del almacenamiento seguro de documentos.")


def archive_catalog_document(path):
    PRODUCT_IMPORT_DIR.mkdir(parents=True, exist_ok=True)
    digest = hashlib.sha256(path.read_bytes()).hexdigest()
    safe_name = re.sub(r"[^A-Za-z0-9._-]+", "-", path.name).strip("-.") or f"catalog{path.suffix.lower()}"
    target = PRODUCT_IMPORT_DIR / f"{digest[:16]}-{safe_name}"
    if not target.exists():
        shutil.copy2(path, target)
    return target, digest


def _existing_product_records():
    records = []
    if not PRODUCT_DIR.exists():
        return records
    for path in sorted(PRODUCT_DIR.glob("*.md")):
        if path.name == "product.example.md":
            continue
        fields = product_fields(read_text(path))
        records.append({"id": path.stem, "path": path, "fields": fields})
    return records


def refresh_catalog_index():
    records = _existing_product_records()
    payload = {
        "schema_version": 1,
        "product_count": len(records),
        "products": [
            {
                "id": item["id"],
                "name": item["fields"].get("name", ""),
                "sku": item["fields"].get("sku", ""),
                "kind": item["fields"].get("kind", ""),
                "category": item["fields"].get("category", ""),
                "price": item["fields"].get("price", ""),
                "tags": item["fields"].get("tags", ""),
                "components": item["fields"].get("components", ""),
                "guide": f"brand_guides/products/{item['path'].name}",
            }
            for item in records
        ],
    }
    CATALOG_INDEX_FILE.parent.mkdir(parents=True, exist_ok=True)
    CATALOG_INDEX_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return payload


def _structured_products(value):
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            return []
    if isinstance(value, dict):
        value = value.get("products") or value.get("productos") or [value]
    return [dict(item) for item in (value or []) if isinstance(item, dict)] if isinstance(value, list) else []


def _canonical_product_record(product):
    source = normalize_product_payload(dict(product or {}))
    canonical = {}
    extras = {}
    for raw_key, raw_value in source.items():
        value = _cell_text(raw_value)
        if not value:
            continue
        key = raw_key if raw_key in HEADER_ALIASES else HEADER_LOOKUP.get(_normalized(raw_key), "")
        if key:
            previous = str(canonical.get(key) or "").strip()
            canonical[key] = f"{previous} | {value}" if previous and value not in previous else value
        elif raw_key not in {"id", "product_id"}:
            extras[str(raw_key)] = value
    if source.get("id"):
        canonical["id"] = str(source.get("id")).strip()
    existing_details = str(canonical.get("additional_details") or "").strip()
    extra_details = _join_extra_fields(extras)
    if extra_details:
        canonical["additional_details"] = " | ".join(item for item in [existing_details, extra_details] if item)
    return canonical


def _dedupe_import_products(products):
    unique = []
    positions = {}
    for raw_product in products:
        product = _canonical_product_record(raw_product)
        name = str(product.get("name") or product.get("product_name") or "").strip()
        sku = str(product.get("sku") or product.get("product_id") or "").strip()
        if not name:
            continue
        key = _normalized(sku) if sku else _normalized(name)
        normalized = dict(product)
        normalized["name"] = name
        if sku:
            normalized["sku"] = sku
        if key in positions:
            unique[positions[key]].update({k: v for k, v in normalized.items() if str(v or "").strip()})
        else:
            positions[key] = len(unique)
            unique.append(normalized)
    return unique


def import_product_catalog(payload):
    payload = dict(payload or {})
    products = _structured_products(payload.get("products") or payload.get("productos"))
    rejected = []
    documents = []
    extracted_texts = []
    notices = []
    raw_paths = payload.get("file_paths") or payload.get("document_paths") or payload.get("file_path") or payload.get("document_path") or []
    if isinstance(raw_paths, str):
        raw_paths = [raw_paths]
    for raw_path in list(raw_paths)[:10]:
        source = safe_catalog_document_path(raw_path)
        archived, digest = archive_catalog_document(source)
        documents.append({"filename": archived.name, "sha256": digest, "path": str(archived)})
        suffix = archived.suffix.lower()
        if suffix in {".csv", ".tsv"}:
            found, invalid = _extract_delimited(archived)
        elif suffix == ".xlsx":
            found, invalid = _extract_xlsx(archived)
        elif suffix == ".xls":
            found, invalid = _extract_xls(archived)
        elif suffix == ".json":
            found, invalid = _extract_json(archived)
        else:
            found, invalid, extracted, notice = _extract_pdf(archived)
            if extracted:
                text_path = archived.with_suffix(archived.suffix + ".extracted.txt")
                text_path.write_text(extracted, encoding="utf-8")
                extracted_texts.append({"source": archived.name, "text_path": str(text_path), "excerpt": extracted[:16000]})
            if notice:
                notices.append(notice)
        products.extend(found)
        rejected.extend({**item, "source": archived.name} for item in invalid)
    products = _dedupe_import_products(products)
    existing = _existing_product_records()
    existing_by_name = {_normalized(item["fields"].get("name")): item for item in existing if item["fields"].get("name")}
    existing_by_sku = {_normalized(item["fields"].get("sku")): item for item in existing if item["fields"].get("sku")}
    incoming_new_keys = set()
    prepared = []
    for product in products:
        name_key = _normalized(product.get("name"))
        sku_key = _normalized(product.get("sku"))
        match = existing_by_sku.get(sku_key) if sku_key else existing_by_name.get(name_key)
        if match:
            product["id"] = match["id"]
        else:
            new_key = sku_key or name_key
            if new_key:
                incoming_new_keys.add(new_key)
        prepared.append(product)
    existing_base = sum(1 for item in existing if _normalized(item["fields"].get("kind")) not in {"bundle", "paquete", "conjunto", "kit"})
    incoming_base_keys = {
        (_normalized(product.get("sku")) or _normalized(product.get("name")))
        for product in prepared
        if _normalized(product.get("kind")) not in {"bundle", "paquete", "conjunto", "kit"}
        and not ((existing_by_sku.get(_normalized(product.get("sku"))) if product.get("sku") else existing_by_name.get(_normalized(product.get("name")))))
    }
    if existing_base + len(incoming_base_keys) > MAX_BASE_PRODUCTS:
        raise ValueError(f"El catálogo admite hasta {MAX_BASE_PRODUCTS} productos base por negocio. Esta importación superaría ese límite.")
    if len(existing) + len(incoming_new_keys) > MAX_PRODUCT_GUIDES:
        raise ValueError(f"El catálogo admite hasta {MAX_PRODUCT_GUIDES} fichas contando productos y ofertas combinadas.")
    saved_ids = []
    for product in prepared:
        if not product.get("source") and documents:
            product["source"] = ", ".join(item["filename"] for item in documents)
        result = save_product_guide(product, refresh=False)
        saved_ids.append(result["product_id"])
    if prepared:
        refresh_offer_map()
    index = refresh_catalog_index()
    needs_structuring = bool(extracted_texts and not prepared)
    return {
        "ok": bool(prepared) or bool(documents),
        "product_count": index["product_count"],
        "imported_count": len(saved_ids),
        "product_ids": saved_ids,
        "documents": documents,
        "rejected_rows": rejected[:100],
        "needs_agent_structuring": needs_structuring,
        "extracted_texts": extracted_texts,
        "notices": notices,
        "capacity": {"base_products": MAX_BASE_PRODUCTS, "total_product_and_offer_guides": MAX_PRODUCT_GUIDES},
        "library": guide_library(),
    }


def search_product_catalog(payload):
    payload = dict(payload or {})
    query = str(payload.get("query") or payload.get("search") or payload.get("product") or "").strip()
    try:
        limit = max(1, min(20, int(payload.get("limit") or 8)))
    except (TypeError, ValueError):
        limit = 8
    query_tokens = set(_normalized(query).split())
    ranked = []
    for item in _existing_product_records():
        fields = item["fields"]
        exact_name = _normalized(fields.get("name"))
        exact_sku = _normalized(fields.get("sku"))
        haystack = " ".join(str(fields.get(key) or "") for key in fields)
        haystack_tokens = set(_normalized(haystack).split())
        score = len(query_tokens & haystack_tokens)
        if query and _normalized(query) in {exact_name, exact_sku, item["id"]}:
            score += 100
        elif query and _normalized(query) in _normalized(haystack):
            score += 10
        if query and score <= 0:
            continue
        ranked.append((score, exact_name, item))
    ranked.sort(key=lambda row: (-row[0], row[1]))
    matches = []
    for score, _, item in ranked[:limit]:
        fields = item["fields"]
        matches.append({
            "id": item["id"],
            "name": fields.get("name", ""),
            "sku": fields.get("sku", ""),
            "kind": fields.get("kind", ""),
            "category": fields.get("category", ""),
            "price": fields.get("price", ""),
            "audience": fields.get("audience", ""),
            "benefit": fields.get("desire", ""),
            "components": fields.get("components", ""),
            "tags": fields.get("tags", ""),
            "guide": f"brand_guides/products/{item['path'].name}",
            "score": score,
        })
    return {
        "ok": True,
        "query": query,
        "total_products": len(_existing_product_records()),
        "matches": matches,
        "capacity": {"base_products": MAX_BASE_PRODUCTS, "total_product_and_offer_guides": MAX_PRODUCT_GUIDES},
    }
